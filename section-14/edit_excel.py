import openpyxl as xl

workbook = xl.Workbook()
sheet_name = workbook.sheetnames[0]
sheet_1 = workbook[sheet_name]

sheet_1['A1'].value = "Hello"
sheet_1['A2'].value = "World"

sheet_2 = workbook.create_sheet(index=0, title="New Sheet Name")
# sheet_2.title = "New Sheet Name"

workbook.save("files/section-14/edit_excel.xlsx")