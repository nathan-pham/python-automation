import openpyxl as xl
# os.chdir("C:\path")

workbook = xl.load_workbook("files/section-14/read_excel.xlsx")
sheets = workbook.sheetnames
sheet = workbook[sheets[0]] # get Sheet1

cell = sheet["B1"]
print(cell.value)

cell = sheet.cell(row=1, column=2)
print(cell.value)